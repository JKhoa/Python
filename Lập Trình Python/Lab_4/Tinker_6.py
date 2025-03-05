from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk, messagebox
import re

wb = load_workbook('C:\\Users\\Admin\\Downloads\\Book2.xlsx')
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 40
    sheet.column_dimensions['H'].width = 50

def clear():
    for field in fields:
        field.delete(0, END)

def validate():
    mssv = mssv_field.get()
    hoten = hoten_field.get()
    ngaysinh = ngaysinh_field.get()
    email = email_field.get()
    sdt = sdt_field.get()
    hocky = hocky_field.get()
    namhoc = namhoc_var.get()

    if not mssv.isdigit() or len(mssv) != 7:
        messagebox.showerror("Lỗi", "Mã số sinh viên phải là 7 chữ số")
        return False

    if not re.match(r"^\w+@[a-zA-Z_]+?\.[a-zA-Z]{2,3}$", email):
        messagebox.showerror("Lỗi", "Email không hợp lệ")
        return False

    if not sdt.isdigit() or len(sdt) != 10:
        messagebox.showerror("Lỗi", "Số điện thoại phải là 10 chữ số")
        return False

    if hocky not in ('1', '2', '3'):
        messagebox.showerror("Lỗi", "Học kỳ chỉ được là 1, 2 hoặc 3")
        return False

    if not re.match(r"^\d{2}/\d{2}/\d{4}$", ngaysinh):
        messagebox.showerror("Lỗi", "Ngày sinh phải đúng định dạng dd/mm/yyyy")
        return False

    if not any(var.get() for var in course_vars):
        messagebox.showerror("Lỗi", "Phải chọn ít nhất một môn học")
        return False

    return True

def insert():
    if validate():
        for i, course in enumerate(courses):
            if course_vars[i].get():
                current_row = sheet.max_row + 1
                sheet.append([mssv_field.get(), hoten_field.get(), ngaysinh_field.get(),
                              email_field.get(), sdt_field.get(), hocky_field.get(),
                              namhoc_var.get(), course])
        wb.save('C:\\Users\\Admin\\Downloads\\Book2.xlsx')
        messagebox.showinfo("Thành công", "Đăng ký thành công")
        clear()

if __name__ == '__main__':
    root = Tk()
    root.configure(background='light green')
    root.title("Đăng ký học phần")
    root.geometry("600x500")

    heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green", fg="red", font=("Arial", 16, "bold"))
    heading.grid(row=0, column=1, columnspan=2, pady=10)

    labels = ["Mã số sinh viên", "Họ tên", "Ngày sinh", "Email", "Số điện thoại", "Học kỳ"]
    fields = []

    for i, label in enumerate(labels):
        lbl = Label(root, text=label, bg="light green")
        lbl.grid(row=i + 1, column=0, padx=10, pady=5, sticky=W)
        entry = Entry(root, width=50)
        entry.grid(row=i + 1, column=1, pady=5)
        fields.append(entry)

    mssv_field, hoten_field, ngaysinh_field, email_field, sdt_field, hocky_field = fields

    namhoc_var = StringVar()
    namhoc_dropdown = ttk.Combobox(root, textvariable=namhoc_var, values=["2022-2023", "2023-2024", "2024-2025"])
    namhoc_dropdown.grid(row=7, column=1, pady=5, sticky=W)

    courses = ["Lập trình Python", "Lập trình Java", "Công nghệ phần mềm", "Phát triển ứng dụng web"]
    course_vars = []

    for i, course in enumerate(courses):
        var = IntVar()
        chk = Checkbutton(root, text=course, bg="light green", variable=var)
        chk.grid(row=8 + i // 2, column=i % 2 + 1, sticky=W, padx=5)
        course_vars.append(var)

    register_btn = Button(root, text="Đăng ký", bg="red", fg="black", command=insert)
    register_btn.grid(row=10, column=1, pady=10, sticky=E, padx=5)

    exit_btn = Button(root, text="Thoát", bg="red", fg="black", command=root.quit)
    exit_btn.grid(row=10, column=2, pady=10, sticky=W)

    root.mainloop()
